#!/usr/bin/env python3
"""
Serveur FastAPI — RenovBat 2026

Lancer depuis la racine du projet :
    uvicorn api.server:app --reload --port 8000

Endpoints :
    GET  /                          → redirige vers /ui/index.html
    GET  /api/budget                → état budgétaire par lot
    POST /api/situations/run        → lance process_situations + generate_recap
    GET  /api/documents             → liste les fichiers générés
    GET  /api/documents/download    → télécharge un fichier (?path=data/...)
    POST /api/reinit                → supprime les fichiers générés
"""

import csv
import json
import pathlib
import subprocess
from datetime import date

import openpyxl
from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles

BASE = pathlib.Path(".")

app = FastAPI(title="RenovBat 2026 API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/ui", StaticFiles(directory="ui"), name="ui")


@app.get("/", include_in_schema=False)
def root():
    return RedirectResponse("/ui/index.html")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _load_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _lots_state() -> dict:
    budgets = {}
    for b in _load_csv("data/budget/budget_lot_prestataire.csv"):
        budgets[b["id_prestataire"]] = {
            "id_lot": b["id_lot"].strip(),
            "budget_total": (float(b["montant_initial"])
                             + float(b["avenant_1"])
                             + float(b["avenant_2"])),
        }

    base_nav = BASE / "data/navettes_et_bons"
    result = {}
    for id_p, bud in budgets.items():
        id_lot = bud["id_lot"]
        search = id_lot.lower().replace("_", " ")
        f = next((x for x in sorted(base_nav.glob("*.xlsx"))
                   if search in x.stem.lower().replace("_", " ")), None)
        engage = 0.0
        if f and f.exists():
            wb = openpyxl.load_workbook(f)
            for sname in wb.sheetnames:
                ws = wb[sname]
                if ws["A18"].value and "APPROUVÉE" in str(ws["A18"].value):
                    try:
                        engage += float(ws["B16"].value or 0)
                    except (TypeError, ValueError):
                        pass
        result[id_lot] = {"budget": bud["budget_total"], "engage": engage}
    return result


# ── GET /api/budget ───────────────────────────────────────────────────────────

@app.get("/api/budget")
def get_budget():
    lots = _lots_state()
    total_budget = sum(v["budget"] for v in lots.values())
    total_engage = sum(v["engage"] for v in lots.values())

    # Dernière session : date du dernier fichier recap
    recap_files = sorted((BASE / "data/vue_globale").glob("recap_*.txt"))
    derniere_session = None
    if recap_files:
        yyyymmdd = recap_files[-1].stem.replace("recap_", "")
        derniere_session = f"{yyyymmdd[6:8]}/{yyyymmdd[4:6]}/{yyyymmdd[:4]}"

    # Nombre de factures approuvées (toutes feuilles de tous les fichiers Excel)
    base_nav = BASE / "data/navettes_et_bons"
    nb_factures = 0
    for xlsx in base_nav.glob("*.xlsx"):
        wb = openpyxl.load_workbook(xlsx)
        for sname in wb.sheetnames:
            ws = wb[sname]
            if ws["A18"].value and "APPROUVÉE" in str(ws["A18"].value):
                nb_factures += 1

    return {
        "lots": lots,
        "total_budget": total_budget,
        "total_engage": total_engage,
        "total_disponible": total_budget - total_engage,
        "nb_factures": nb_factures,
        "derniere_session": derniere_session,
    }


# ── POST /api/situations/run ──────────────────────────────────────────────────

@app.post("/api/situations/run")
def run_situations():
    r1 = subprocess.run(
        ["python3", "scripts/process_situations.py"],
        capture_output=True, text=True,
    )
    if r1.returncode != 0:
        raise HTTPException(status_code=500, detail=r1.stderr)

    r2 = subprocess.run(
        ["python3", "scripts/generate_recap.py"],
        capture_output=True, text=True,
    )
    if r2.returncode != 0:
        raise HTTPException(status_code=500, detail=r2.stderr)

    session_file = BASE / "data/tmp/.current_session.json"
    factures = []
    if session_file.exists():
        factures = json.loads(session_file.read_text())["factures_session"]

    return {"factures": factures, "lots": _lots_state()}


# ── GET /api/documents ────────────────────────────────────────────────────────

@app.get("/api/documents")
def list_documents():
    nav = BASE / "data/navettes_et_bons"
    vue = BASE / "data/vue_globale"

    def info(f: pathlib.Path):
        return {
            "name": f.name,
            "path": str(f.relative_to(BASE)),
            "date": date.fromtimestamp(f.stat().st_mtime).strftime("%d/%m/%Y"),
        }

    return {
        "navettes": [info(f) for f in sorted(nav.glob("*.xlsx"))],
        "mails": [info(f) for f in sorted((nav / "mails").glob("mail_*.txt"))],
        "vue_globale": [info(f) for f in sorted(vue.iterdir())
                        if f.suffix in (".png", ".xlsx", ".txt")
                        and not f.name.startswith(".")],
    }


# ── GET /api/documents/download ───────────────────────────────────────────────

@app.get("/api/documents/download")
def download_document(path: str = Query(...)):
    f = (BASE / path).resolve()
    data_dir = (BASE / "data").resolve()
    if not f.is_file() or not str(f).startswith(str(data_dir)):
        raise HTTPException(status_code=403)
    return FileResponse(f, filename=f.name)


# ── POST /api/reinit ──────────────────────────────────────────────────────────

@app.post("/api/reinit")
def reinit():
    base = BASE / "data"
    patterns = [
        "navettes_et_bons/*.xlsx",
        "navettes_et_bons/mails/mail_*.txt",
        "navettes_et_bons/rejets/rejet_*.txt",
        "vue_globale/recap_*.txt",
        "vue_globale/budget_*.png",
        "vue_globale/budget_*.xlsx",
        "tmp/.current_session.json",
    ]
    deleted = []
    for p in patterns:
        for f in base.glob(p):
            f.unlink()
            deleted.append(f.name)
    return {"deleted": deleted, "count": len(deleted)}
