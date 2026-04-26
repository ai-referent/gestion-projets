#!/usr/bin/env python3
"""
Traitement des factures PDF nouvelles — Projet renovation_2026
Étapes 1 à 5 du workflow situations.

- Charge prestataires.csv et budget_lot_prestataire.csv
- Parse chaque PDF dans data/factures/
- Identifie le prestataire (SIRET en priorité, sinon nom approché)
- Vérifie le budget et appelle scripts/generate_navette.py
- Génère les mails de transmission dans data/navettes_et_bons/mails/
- Écrit data/vue_globale/.current_session.json pour generate_recap.py
"""

import csv
import json
import pathlib
import re
import subprocess
import unicodedata
import xml.etree.ElementTree as ET
from datetime import date

import openpyxl

# ── Étape 1 — Charger les données de référence ────────────────────────────────

def _load_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))

prestataires = {p["id_prestataire"]: p
                for p in _load_csv("data/prestataires/prestataires.csv")}

budgets = {}
lot_nums = {}
for _i, _b in enumerate(_load_csv("data/budget/budget_lot_prestataire.csv"), start=1):
    _id = _b["id_prestataire"]
    budgets[_id] = {
        "id_lot": _b["id_lot"].strip(),
        "budget_total": (float(_b["montant_initial"])
                         + float(_b["avenant_1"])
                         + float(_b["avenant_2"])),
    }
    lot_nums[_id] = _i

factures_pdf = sorted(pathlib.Path("data/factures").glob("*.pdf"))

# ── Étape 2 — Parser les PDF ───────────────────────────────────────────────────

_NS = {
    "rsm": "urn:un:unece:uncefact:data:standard:CrossIndustryInvoice:100",
    "ram": "urn:un:unece:uncefact:data:standard:ReusableAggregateBusinessInformationEntity:100",
    "udt": "urn:un:unece:uncefact:data:standard:UnqualifiedDataType:100",
}
_MOIS = ["", "janvier", "fevrier", "mars", "avril", "mai", "juin",
         "juillet", "aout", "septembre", "octobre", "novembre", "decembre"]


def _parse_xml(data):
    m = re.search(b"<rsm:CrossIndustryInvoice.*?</rsm:CrossIndustryInvoice>",
                  data, re.DOTALL)
    if not m:
        return None
    try:
        root = ET.fromstring(m.group(0).decode("utf-8"))
    except ET.ParseError:
        return None
    inv_id = root.findtext("rsm:ExchangedDocument/ram:ID", namespaces=_NS)
    d      = root.findtext("rsm:ExchangedDocument/ram:IssueDateTime/udt:DateTimeString",
                           namespaces=_NS)
    seller = root.findtext(".//ram:SellerTradeParty/ram:Name", namespaces=_NS)
    total  = root.findtext(".//ram:GrandTotalAmount", namespaces=_NS)
    siret  = root.findtext(
        ".//ram:SellerTradeParty/ram:SpecifiedLegalOrganization/ram:ID",
        namespaces=_NS)
    if not all([inv_id, d, seller, total]):
        return None
    return {
        "numero": inv_id.strip(),
        "date": f"{int(d[6:8])} {_MOIS[int(d[4:6])]} {d[:4]}",
        "societe": seller.strip(),
        "montant_ttc": float(total.strip()),
        "siret": siret.strip() if siret else None,
    }


def _parse_streams(data):
    texts = []
    for s in re.findall(b"stream\n(.*?)\nendstream", data, re.DOTALL):
        for line in s.decode("latin-1").split("\n"):
            m = re.search(r"\(([^)]+)\)", line)
            if m:
                texts.append(m.group(1).strip())
    fac = next((t for t in texts if "FAC-" in t), None)
    dat = next((t for t in texts if "Date" in t), None)
    soc = next((t for t in texts if len(t) > 5
                and "avenue" not in t.lower()
                and "SIRET" not in t and "FAC-" not in t
                and "Date" not in t and "Client" not in t), None)
    idx = next((i for i, t in enumerate(texts) if "TOTAL TTC" in t), None)
    if not all([fac, dat, soc, idx is not None]):
        return None
    amt = re.sub(r"[^\d,.]", "",
                 texts[idx + 1] if idx + 1 < len(texts) else "").replace(",", ".")
    try:
        return {
            "numero": re.sub(r"^[^:]*:\s*", "", fac).strip(),
            "date": re.sub(r"^[Dd]ate[^:]*:\s*", "", dat).strip(),
            "societe": soc,
            "montant_ttc": float(amt),
            "siret": None,
        }
    except ValueError:
        return None


def parse_facture(path):
    data = pathlib.Path(path).read_bytes()
    return _parse_xml(data) or _parse_streams(data)


# ── Étape 3 — Identifier le prestataire ───────────────────────────────────────

def _normalize(s):
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", s.lower())


def find_prestataire(fac):
    if fac.get("siret"):
        for id_p, p in prestataires.items():
            if p["siret"] == fac["siret"]:
                return id_p
    fn = _normalize(fac["societe"])
    for id_p, p in prestataires.items():
        pn = _normalize(p["nom_prestataire"])
        if pn in fn or fn in pn:
            return id_p
    return None


# ── Boucle principale ──────────────────────────────────────────────────────────

base_nav = pathlib.Path("data/navettes_et_bons")
mails_dir  = base_nav / "mails"
rejets_dir = base_nav / "rejets"
mails_dir.mkdir(parents=True, exist_ok=True)
rejets_dir.mkdir(parents=True, exist_ok=True)
tmp_dir = pathlib.Path("data/tmp")
tmp_dir.mkdir(parents=True, exist_ok=True)
today_str = date.today().strftime("%d/%m/%Y")


def _write_rejet(numero, societe, lot, montant, motif):
    contenu = f"""\
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
✗ REJET — {numero}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Date     : {today_str}
Société  : {societe}
Lot      : {lot}
Montant  : {montant:,.2f} EUR
Motif    : {motif}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
    (rejets_dir / f"rejet_{numero}.txt").write_text(contenu, encoding="utf-8")
    print(f"  → Rejet enregistré : rejet_{numero}.txt")

factures_session = []

for pdf_path in factures_pdf:
    print(f"\n── {pdf_path.name} ──")

    fac = parse_facture(pdf_path)
    if not fac:
        print("  ✗ Impossible de parser ce PDF.")
        continue

    # Identification du prestataire
    id_p = find_prestataire(fac)
    if not id_p:
        motif_rejet = "Société non référencée sur le projet"
        print(f"  ✗ REJET : {motif_rejet} ({fac['societe']})")
        _write_rejet(fac["numero"], fac["societe"], "—", fac["montant_ttc"], motif_rejet)
        factures_session.append({
            "ref": fac["numero"], "societe": fac["societe"],
            "lot": "—", "ttc": fac["montant_ttc"],
            "statut": "Rejetée",
            "motif": motif_rejet,
        })
        continue

    id_lot = budgets[id_p]["id_lot"]
    budget_total = budgets[id_p]["budget_total"]

    # Localiser le fichier Excel du lot
    search = id_lot.lower().replace("_", " ")
    target_file = None
    for f in sorted(base_nav.glob("*.xlsx")):
        if search in f.stem.lower().replace("_", " "):
            target_file = f
            break
    if not target_file:
        num = lot_nums[id_p]
        cap = id_lot.replace("_", " ").title().replace(" ", "_")
        target_file = base_nav / f"lot{num:02d}_{cap}_{id_p}.xlsx"

    # Vérifier si la facture est déjà traitée
    if target_file.exists():
        wb = openpyxl.load_workbook(target_file)
        if fac["numero"] in wb.sheetnames:
            print(f"  → {fac['numero']} : déjà traitée, ignorée.")
            continue
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Calculer le cumul des factures approuvées existantes
    cumul_existant = 0.0
    for sname in wb.sheetnames:
        ws_ex = wb[sname]
        if ws_ex["A18"].value and "APPROUVÉE" in str(ws_ex["A18"].value):
            try:
                cumul_existant += float(ws_ex["B16"].value or 0)
            except (TypeError, ValueError):
                pass

    # Vérification budgétaire
    montant = fac["montant_ttc"]
    if cumul_existant + montant > budget_total:
        approved = False
        motif = (f"Dépassement du budget lot "
                 f"(budget={budget_total:.2f}€, "
                 f"engagé={cumul_existant:.2f}€, "
                 f"facture={montant:.2f}€)")
    else:
        approved = True
        motif = None

    # Étape 4 — Appeler generate_navette.py
    payload = {
        "target_file":     str(target_file),
        "sheet_name":      fac["numero"],
        "fac": {
            "numero":      fac["numero"],
            "date":        fac["date"],
            "societe":     fac["societe"],
            "montant_ttc": montant,
        },
        "id_lot":          id_lot,
        "lot_num":         lot_nums[id_p],
        "adresse":         prestataires[id_p].get("adresse_prestataire", ""),
        "approved":        approved,
        "motif":           motif,
        "cumul_existant":  cumul_existant,
        "budget_total":    budget_total,
    }
    result = subprocess.run(
        ["python3", "scripts/generate_navette.py", json.dumps(payload)],
        capture_output=True, text=True,
    )
    print("  →", result.stdout.strip())
    if result.returncode != 0:
        print("  ERREUR:", result.stderr.strip())
        continue

    if not approved:
        _write_rejet(fac["numero"], fac["societe"], id_lot, montant, motif)

    # Étape 5 — Générer le mail de transmission (si approuvée)
    if approved:
        nouveau_cumul = cumul_existant + montant
        contenu = f"""\
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📨 MAIL — TRANSMISSION RenovBat → IMMOSOCIAL_69
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
De     : RenovBat (MOE) <moe@renovbat.fr>
À      : IMMOSOCIAL_69 (MOA) <moa@immosocial69.fr>
Date   : {today_str}
Objet  : Fiche navette et bon de paiement — Facture {fac['numero']}

Madame, Monsieur,

Veuillez trouver ci-joint la fiche navette et le bon de paiement
relatifs à la facture n° {fac['numero']} du {fac['date']},
émise par {fac['societe']} dans le cadre du lot {id_lot}
du projet renovation_2026.

Statut            : ✓ APPROUVÉE
Montant TTC       : {montant:,.2f} EUR
Nouveau cumul lot : {nouveau_cumul:,.2f} EUR
Reste à régler    : {budget_total - nouveau_cumul:,.2f} EUR

Cordialement,
RenovBat — Maître d'Œuvre du projet renovation_2026
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""
        (mails_dir / f"mail_{fac['numero']}.txt").write_text(contenu, encoding="utf-8")
        print(f"  → Mail généré : mail_{fac['numero']}.txt")

    factures_session.append({
        "ref":    fac["numero"],
        "societe": fac["societe"],
        "lot":    id_lot,
        "ttc":    montant,
        "statut": "Approuvée" if approved else "Rejetée",
    })

# Sauvegarder pour generate_recap.py
(tmp_dir / ".current_session.json").write_text(
    json.dumps({"factures_session": factures_session}, ensure_ascii=False, indent=2),
    encoding="utf-8",
)

print(f"\n✓ {len(factures_session)} facture(s) traitée(s) dans cette session.")
