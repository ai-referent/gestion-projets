#!/usr/bin/env python3
"""Génère ou met à jour une feuille fiche-navette + bon de paiement dans un classeur xlsx.

Usage:
    python3 scripts/generate_navette.py '<json>'

JSON attendu:
{
  "target_file"   : "data/navettes_et_bons/lot01_...",
  "sheet_name"    : "FAC-2026-001",
  "fac"           : {"numero": "FAC-2026-001", "date": "15 mars 2026",
                     "societe": "Platr'O'Matic SARL", "montant_ttc": 12500.0},
  "id_lot"        : "isolation_thermique",
  "lot_num"       : 1,
  "adresse"       : "42 rue du Crepissage - 75011 PARIS",
  "approved"      : true,
  "motif"         : null,
  "cumul_existant": 0.0,
  "budget_total"  : 20000.0
}

Cellules de référence pour la lecture par process_situations.py et generate_recap.py :
  B14 = montant HT, B15 = TVA, B16 = montant TTC
  A18 = marqueur approbation ("✓ FACTURE APPROUVÉE" ou "✗ FACTURE REJETÉE")

Sortie stdout: message de statut.
"""

import json
import pathlib
import sys
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

TVA_TAUX = 0.20

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _eur(cell):
    cell.number_format = '#,##0.00 "EUR"'


def _green_banner(ws, cell_ref, merge):
    ws[cell_ref] = "RenovBat Bureau d'Etude Bâtiment"
    ws[cell_ref].font = Font(bold=True, color="FFFFFF")
    ws[cell_ref].fill = PatternFill("solid", fgColor="1E8449")
    ws[cell_ref].alignment = Alignment(horizontal="center")
    ws.merge_cells(merge)


def add_sheet(data: dict) -> str:
    target_file = pathlib.Path(data["target_file"])
    sheet_name  = data["sheet_name"]
    fac         = data["fac"]
    id_lot      = data["id_lot"]
    lot_num     = data.get("lot_num", "")
    adresse     = data.get("adresse", "")
    approved    = data["approved"]
    motif       = data.get("motif")
    cumul_existant_ttc = float(data.get("cumul_existant", 0))
    budget_ht          = float(data.get("budget_total", 0))

    montant_ttc = float(fac["montant_ttc"])
    montant_ht  = round(montant_ttc / (1 + TVA_TAUX), 2)
    montant_tva = round(montant_ttc - montant_ht, 2)
    budget_ttc  = round(budget_ht * (1 + TVA_TAUX), 2)
    today       = date.today().strftime("%d/%m/%Y")

    lot_display = f"lot{lot_num} : {id_lot.replace('_', ' ').capitalize()}"

    # Charger ou créer le classeur
    if target_file.exists():
        wb = openpyxl.load_workbook(target_file)
        if sheet_name in wb.sheetnames:
            return f"IGNORÉE (déjà traitée) — {target_file.name}"
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws = wb.create_sheet(title=sheet_name)

    GREEN = "1E8449"
    RED   = "C0392B"
    BOLD  = Font(bold=True)

    # ── EN-TÊTE TITULAIRE (lignes 1-4) ───────────────────────────────────────

    for col, label in zip("ABCD", ["ENTREPRISE", "lot", "sit n°", "fiche navette n°"]):
        ws[f"{col}1"] = label
        ws[f"{col}1"].font = BOLD
        ws[f"{col}1"].border = _BORDER

    ws["A2"] = fac["societe"];   ws["A2"].font = BOLD; ws["A2"].border = _BORDER
    ws["B2"] = lot_display;      ws["B2"].font = BOLD; ws["B2"].border = _BORDER
    ws["C2"].border = _BORDER
    ws["D2"].border = _BORDER

    ws["A3"] = adresse;          ws["A3"].font = BOLD

    # ── FICHE NAVETTE (lignes 5-20) ──────────────────────────────────────────

    ws["A5"] = "FICHE NAVETTE — PROJET renovation_2026"
    ws["A5"].font = Font(bold=True, size=13)
    ws.merge_cells("A5:B5")

    _green_banner(ws, "A6", "A6:B6")

    ws["A8"] = "MOA";  ws["B8"] = "IMMOSOCIAL_69"

    ws["A10"] = "Référence facture";  ws["B10"] = fac["numero"]
    ws["A11"] = "Date de la facture"; ws["B11"] = fac["date"]
    ws["A12"] = "Émetteur";           ws["B12"] = fac["societe"]
    ws["A13"] = "Lot concerné";       ws["B13"] = id_lot

    # Détail HT / TVA / TTC  (B14, B15, B16 = cellules de référence)
    ws["A14"] = "Montant HT";             ws["B14"] = montant_ht;  _eur(ws["B14"])
    ws["A15"] = f"TVA ({TVA_TAUX:.0%})";  ws["B15"] = montant_tva; _eur(ws["B15"])
    ws["A16"] = "Montant TTC";            ws["B16"] = montant_ttc; _eur(ws["B16"])

    # Statut — A18 est la cellule de référence lue par process_situations.py
    if approved:
        ws["A18"] = "✓ FACTURE APPROUVÉE"
        ws["A18"].font = Font(bold=True, color=GREEN)
        ws.merge_cells("A18:B18")
        ws["A19"] = "Date d'approbation"; ws["B19"] = today
        ws["A20"] = "Approuvé par";       ws["B20"] = "RenovBat"
    else:
        ws["A18"] = "✗ FACTURE REJETÉE"
        ws["A18"].font = Font(bold=True, color=RED)
        ws.merge_cells("A18:B18")
        ws["A19"] = "Date de traitement"; ws["B19"] = today
        ws["A20"] = "Motif du rejet";     ws["B20"] = motif or ""
        ws["A21"] = "Traité par";         ws["B21"] = "RenovBat"

    # ── BON DE PAIEMENT (lignes 23-40, uniquement si approuvée) ──────────────
    if approved:
        nouveau_cumul_ttc = cumul_existant_ttc + montant_ttc

        ws["A23"] = "BON DE PAIEMENT — PROJET renovation_2026"
        ws["A23"].font = Font(bold=True, size=13)
        ws.merge_cells("A23:B23")

        _green_banner(ws, "A24", "A24:B24")

        ws["A26"] = "Budget total du lot HT";  ws["B26"] = budget_ht;  _eur(ws["B26"])
        ws["A27"] = "Budget total du lot TTC"; ws["B27"] = budget_ttc; _eur(ws["B27"])

        ws["A29"] = "Situations précédentes TTC"; ws["B29"] = cumul_existant_ttc; _eur(ws["B29"])
        ws["A30"] = "Présente situation HT";      ws["B30"] = montant_ht;         _eur(ws["B30"])
        ws["A31"] = "Présente situation TVA";     ws["B31"] = montant_tva;        _eur(ws["B31"])
        ws["A32"] = "Présente situation TTC";     ws["B32"] = montant_ttc;        _eur(ws["B32"])

        ws["A34"] = "Nouveau cumul TTC"
        ws["B34"] = nouveau_cumul_ttc
        ws["B34"].font = Font(bold=True)
        _eur(ws["B34"])

        ws["A35"] = "Reste à régler TTC"
        ws["B35"] = round(budget_ttc - nouveau_cumul_ttc, 2)
        _eur(ws["B35"])

        ws["A37"] = "Date d'émission";  ws["B37"] = today
        ws["A38"] = "Établi par";       ws["B38"] = "RenovBat (MOE)"
        ws["A39"] = "Signataire";       ws["B39"] = "J. Pons"
        ws["A40"] = "À destination de"; ws["B40"] = "IMMOSOCIAL_69 (MOA)"

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18

    target_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_file)
    statut = "APPROUVÉE ✓" if approved else "REJETÉE ✗"
    return f"{sheet_name} : {statut} — {target_file.name}"


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: generate_navette.py '<json>'", file=sys.stderr)
        sys.exit(1)
    result = add_sheet(json.loads(sys.argv[1]))
    print(result)
