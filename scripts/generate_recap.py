#!/usr/bin/env python3
"""
Génère le récapitulatif global — Projet renovation_2026
Étape 6 du workflow situations.

Lit data/vue_globale/.current_session.json produit par process_situations.py.
Produit dans data/vue_globale/ :
  - recap_<AAAAMMJJ>.txt
  - budget_<AAAAMMJJ>.png
  - budget_<AAAAMMJJ>.xlsx
"""

import csv
import json
import pathlib
from datetime import date

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

# ── Charger les données de session ────────────────────────────────────────────

vue_dir = pathlib.Path("data/vue_globale")
session_path = vue_dir / ".current_session.json"
if not session_path.exists():
    raise SystemExit("Aucune session en cours. Lancer d'abord process_situations.py.")

session_data = json.loads(session_path.read_text(encoding="utf-8"))
factures_session = session_data["factures_session"]

# ── Charger le budget de référence ────────────────────────────────────────────

def _load_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))

budgets = {}
for b in _load_csv("data/budget/budget_lot_prestataire.csv"):
    budgets[b["id_prestataire"]] = {
        "id_lot": b["id_lot"].strip(),
        "budget_total": (float(b["montant_initial"])
                         + float(b["avenant_1"])
                         + float(b["avenant_2"])),
    }

# ── Recalculer lots_state depuis les fichiers Excel ────────────────────────────

base_nav = pathlib.Path("data/navettes_et_bons")
lots_state = {}
for id_p, bud in budgets.items():
    id_lot = bud["id_lot"]
    search = id_lot.lower().replace("_", " ")
    f = next((x for x in sorted(base_nav.glob("*.xlsx"))
               if search in x.stem.lower().replace("_", " ")), None)
    engage = 0.0
    if f and f.exists():
        wb_r = openpyxl.load_workbook(f)
        for sname in wb_r.sheetnames:
            ws_r = wb_r[sname]
            if ws_r["A14"].value and "APPROUVÉE" in str(ws_r["A14"].value):
                try:
                    engage += float(ws_r["B12"].value or 0)
                except (TypeError, ValueError):
                    pass
    lots_state[id_lot] = {"budget": bud["budget_total"], "engage": engage}

# ── Dates ──────────────────────────────────────────────────────────────────────

today = date.today()
yyyymmdd = today.strftime("%Y%m%d")
ddmmyyyy = today.strftime("%d/%m/%Y")
vue_dir.mkdir(parents=True, exist_ok=True)

# ── a) Récapitulatif texte ────────────────────────────────────────────────────

header = ["Référence", "Société", "Lot", "Montant TTC (€)", "Statut"]
rows = [
    [f["ref"], f["societe"], f["lot"],
     f"{f['ttc']:,.2f}".replace(",", " "),
     "✓ Approuvée" if f["statut"] == "Approuvée" else "✗ Rejetée"]
    for f in factures_session
]
col_w = [max(len(str(r[i])) for r in [header] + rows) for i in range(5)]

def _fmt_row(row):
    return "| " + " | ".join(str(v).ljust(col_w[i]) for i, v in enumerate(row)) + " |"

sep = "|-" + "-|-".join("-" * w for w in col_w) + "-|"
lines_fac = [_fmt_row(header), sep] + [_fmt_row(r) for r in rows]

lines_bud = []
for lot, v in lots_state.items():
    pct = v["engage"] / v["budget"] * 100 if v["budget"] > 0 else 0
    lines_bud.append(
        f"{lot:<26}: {v['engage']:>10,.0f} € / {v['budget']:>10,.0f} € ({pct:.1f} %)"
        .replace(",", " ")
    )

recap_content = "\n".join([
    "=" * 44,
    "RÉCAPITULATIF — PROJET renovation_2026",
    "=" * 44,
    f"Date : {ddmmyyyy}",
    "",
    "--- Factures traitées ---",
    *lines_fac,
    "",
    "--- Consommation budgétaire par lot ---",
    *lines_bud,
    "=" * 44,
])
(vue_dir / f"recap_{yyyymmdd}.txt").write_text(recap_content, encoding="utf-8")
print(f"✓ Récapitulatif : recap_{yyyymmdd}.txt")

# ── b) Graphique PNG ───────────────────────────────────────────────────────────

labels    = list(lots_state.keys())
budgets_v = [lots_state[l]["budget"] for l in labels]
engages   = [lots_state[l]["engage"] for l in labels]
pcts      = [e / b * 100 if b > 0 else 0 for e, b in zip(engages, budgets_v)]
colors    = ["#e74c3c" if p > 90 else "#e67e22" if p > 70 else "#2ecc71" for p in pcts]

fig, ax = plt.subplots(figsize=(9, 4))
y = range(len(labels))
ax.barh(y, engages, color=colors, height=0.5, label="Engagé")
ax.barh(y, [b - e for b, e in zip(budgets_v, engages)],
        left=engages, color="#ecf0f1", height=0.5, edgecolor="#bdc3c7", label="Disponible")
for i, (e, b, p) in enumerate(zip(engages, budgets_v, pcts)):
    ax.text(b + 500, i, f"{p:.1f} %  ({e:,.0f} / {b:,.0f} €)".replace(",", " "),
            va="center", fontsize=9)
ax.set_yticks(list(y))
ax.set_yticklabels([l.replace("_", " ").title() for l in labels])
ax.set_xlabel("Montant (EUR)")
ax.set_title("Consommation budgétaire — Projet renovation_2026",
             fontsize=12, fontweight="bold")
ax.set_xlim(0, max(budgets_v) * 1.45)
ax.legend(loc="lower right", fontsize=8)
ax.grid(axis="x", linestyle="--", alpha=0.4)
fig.tight_layout()
fig.savefig(vue_dir / f"budget_{yyyymmdd}.png", dpi=130)
plt.close()
print(f"✓ Graphique     : budget_{yyyymmdd}.png")

# ── c) Classeur Excel ──────────────────────────────────────────────────────────

wb_out = openpyxl.Workbook()

ws1 = wb_out.active
ws1.title = "Factures"
ws1.append(["Référence", "Société", "Lot", "Montant TTC (€)", "Statut"])
for cell in ws1[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2C3E50")
    cell.alignment = Alignment(horizontal="center")
for f in factures_session:
    ws1.append([f["ref"], f["societe"], f["lot"], f["ttc"],
                "✓ Approuvée" if f["statut"] == "Approuvée" else "✗ Rejetée"])
for col, w in zip("ABCDE", [16, 26, 24, 18, 14]):
    ws1.column_dimensions[col].width = w

ws2 = wb_out.create_sheet("Budget")
ws2.append(["Lot", "Budget (€)", "Engagé (€)", "Disponible (€)", "% consommé"])
for cell in ws2[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2C3E50")
for lot_name, v in lots_state.items():
    ws2.append([
        lot_name, v["budget"], v["engage"],
        v["budget"] - v["engage"],
        round(v["engage"] / v["budget"] * 100, 1) if v["budget"] > 0 else 0,
    ])
for col, w in zip("ABCDE", [26, 16, 16, 16, 14]):
    ws2.column_dimensions[col].width = w

chart = BarChart()
chart.type = "bar"
chart.grouping = "stacked"
chart.title = "Budget vs Engagé par lot"
chart.width, chart.height = 18, 10
data_ref = Reference(ws2, min_col=2, max_col=4, min_row=1, max_row=1 + len(lots_state))
cats = Reference(ws2, min_col=1, min_row=2, max_row=1 + len(lots_state))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats)
ws2.add_chart(chart, "F2")

wb_out.save(vue_dir / f"budget_{yyyymmdd}.xlsx")
print(f"✓ Classeur      : budget_{yyyymmdd}.xlsx")

# ── Affichage final ────────────────────────────────────────────────────────────
print()
print(recap_content)
